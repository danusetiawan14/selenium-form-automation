from openpyxl import load_workbook

from src.config import INPUT_FILE


def read_excel():
    """
    Membaca data dari file Excel.
    """

    wb = load_workbook(INPUT_FILE)

    ws = wb.active

    data = []

    for row in ws.iter_rows(min_row=2, values_only=True):

        customer = {
            "first_name": row[0],
            "last_name": row[1],
            "email": row[2],
        }

        data.append(customer)

    return data